# -*- coding: utf-8 -*-
''' Python Spider grab the exchange data '''
### developed by Jansen Liu
# import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup

import time
# from datetime import datetime
import datetime
import holidays
from holidays import countries
import calendar

import re
import mysql.connector
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment
# from openpyxl import load_workbook

### Setting the mysql database connection
def mysql_connection():
    conn = mysql.connector.connect(
        host = "localhost",   # db host
        user = "root",          # db username
        password = "123456",    # db password
        database = "test"       # db name
    )
    cursor = conn.cursor()

    return cursor, conn


### check today's date is in database or not
def check_data_in_db(date):

    cursor, conn = mysql_connection()
    sql = "SELECT DISTINCT 'date' FROM t_forex_exch"

    try:
        cursor.execute(sql)
        # print("Connected to the OA_DB to select the date from t_forex_exch table")
        results = cursor.fetchall()
    except mysql.connector.Error as error:
        print("Failed to select date from t_forex_exch table {}".format(error))
    datestr = [i[0] for i in results]
    
    return str(date) not in datestr

### check the href is in database or not
def check_href_in_db(href):

    cursor, conn = mysql_connection()
    sql = "select distinct href from t_forex_exch"

    try:
        cursor.execute(sql)
        # print("Connected to the OA_DB to select the href from t_forex_exch table")
        results = cursor.fetchall()
    except mysql.connector.Error as error:
        print("Failed to select href from t_forex_exch table {}".format(error))

    hreflist = [s[0] for s in results]

    return str(href) not in hreflist
    

### get the url about exchange notice webpage
def get_html_url(options, url):
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    html_content = driver.page_source
    soup = BeautifulSoup(html_content, 'html.parser')
    atags = soup.select('a')

    anew = None
    for atag in atags:
        if atag.get_text().endswith('中国外汇交易中心受权公布人民币汇率中间价公告'):
            # print(f"{tag.get_text()}: http://www.pbc.gov.cn{tag.get('href')}")
            anew = 'http://www.pbc.gov.cn' + atag.get('href')
            break
    driver.quit()

    return str(anew)


### check if the url has been updated
def check_update(old_content, new_content):
    return old_content != new_content


### get the foreign exchange information
def data_obtain(driver, url):
    driver.get(url)
    url_source = driver.page_source

    soup = BeautifulSoup(url_source, 'html.parser')
    p_tags = soup.select('p')

    notice = ""
    for ptag in p_tags:
        if ptag.get_text().startswith('中国人民银行授权中国外汇交易中心公布'):
            notice = ptag.get_text()

    sentences = []
    lenSentence = []
    for sentence in notice.split("，"):
        if '人民币' in sentence:
            lenSentence.append(len(sentence))
            if len(sentence) > 19:
                s = sentence[-14:-1] + '元'
                sentences.append(s)
            else:
                sentences.append(sentence)
    
    c_to_r = []
    r_to_c = []
    for i in sentences:
        if i[0].isdigit():
            c_to_r.append(i)
        else:
            r_to_c.append(i)

    print(f'外币对人民币: \n{c_to_r}')
    print(f'人民币对外币: \n{r_to_c}')
    return c_to_r, r_to_c


### process the information and save it
def process_save(list1, list2, date, href):

    today = date
    cursor, conn = mysql_connection()

    ### define the data insertion SQL
    insert_sql = "INSERT INTO t_forex_exch(date, forex_rate, currency, rmb_rate, href, status_sent, status_attachment) VALUES(%s, %s, %s, %s, %s, %s, %s)"
    status_sent = '0'
    status_attachment = '0'

    for c in list1:
        matchn = re.search(r'(\w+)\s*对\s*人民币(\d+.\d+)', c)
        c_value = matchn.group(1)
        r_value = "{:.4f}".format(float(matchn.group(2)))
        v = re.findall(r'\d+|[^\d\s]+', c_value)

        # if v[1] == '日元':
        #     v[0] = int(v[0]) / 100
        #     r_value = round(r_value / 100, 4)

        print(f'{v[1]} : {v[0]} : {r_value}')

        ### insert the data into database
        try:
            cursor.execute(insert_sql, (today, v[0], v[1], r_value, href, status_sent, status_attachment))
            conn.commit()
            print(f'New Forex data insert into database succeed!')
        except mysql.connector.Error as error:
            ### output the error information
            print(f'Error inserting data: {error}')


    for r in list2:
        matchr = re.search(r'人民币\s*(\d+)元对\s*([\d.]+)\s*([\w\s]+)', r)
        r1 = matchr.group(1)
        c1 = "{:.4f}".format(float(matchr.group(2)))
        cname = matchr.group(3)
        print(f'{cname} : {c1} : {r1}')

        try:
            cursor.execute(insert_sql, (today, c1, cname, r1, href, status_sent, status_attachment))
            conn.commit()
            print(f'New Forex data insert into database succeed!')
        except mysql.connector.Error as error:
            ### output the error information
            print(f'Error inserting data: {error}')

    ### close the connection of database
    cursor.close()
    conn.close()

    ### write the exchange information into txt file
    # with open('output.txt', 'a') as f:
    #     f.write('---------------------------\n')
    #     f.write(f'{today}: \n')
    #     for item in data:
    #         f.write(f'{item}\n')
    #     f.write('---------------------------\n')
    return


### AUTO Save the forex data into excel table "forex_table" on the end of each month
def excel_forex_save(date, checktime):

    # connect to the "t_forex_exch" database and select the forex data convert to the forex table format
    cursor, conn = mysql_connection()
    part1 = """SELECT DISTINCT fe.date,
                 COALESCE(u.rmb_rate, 0) AS '1usd->rmb_rate',  
                 ROUND(1/COALESCE(u.rmb_rate, 0), 4) AS '1rmb_rate->usd',  
                 COALESCE(h.rmb_rate, 0) AS '1hkd->rmb_rate',  
                 ROUND(1/COALESCE(h.rmb_rate, 0), 4) AS '1rmb_rate->hkd',  
                 ROUND(COALESCE(j.rmb_rate, 0)/100, 4) AS '1jpy->rmb_rate',  
                 ROUND(1/(COALESCE(j.rmb_rate, 0)/100), 4) AS '1rmb_rate->jpy',  
                 ROUND(COALESCE(h.rmb_rate, 0)/COALESCE(u.rmb_rate, 0), 4) AS '1hkd->usd',  
                 ROUND(COALESCE(u.rmb_rate, 0)/COALESCE(h.rmb_rate, 0), 4) AS '1usd->hkd',  
                 ROUND(ROUND(COALESCE(j.rmb_rate, 0)/100, 4)/COALESCE(u.rmb_rate, 0), 4) AS '1jpy->usd',  
                 ROUND(COALESCE(u.rmb_rate, 0)*ROUND(1/(COALESCE(j.rmb_rate, 0)/100), 4), 4) AS '1usd->jpy',  
                 COALESCE(e.rmb_rate, 0) AS '1eur->rmb_rate',  
                 ROUND(1/COALESCE(e.rmb_rate, 0), 4) AS '1rmb_rate->eur'"""

    part2 = """FROM t_forex_exch AS fe  
               LEFT JOIN t_forex_exch AS u ON fe.date = u.date AND u.currency = '美元'  
               LEFT JOIN t_forex_exch AS j ON fe.date = j.date AND j.currency = '日元'  
               LEFT JOIN t_forex_exch AS h ON fe.date = h.date AND h.currency = '港元'  
               LEFT JOIN t_forex_exch AS e ON fe.date = e.date AND e.currency = '欧元'"""

    part3 = """WHERE 
                DATE_FORMAT(fe.date, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m') AND  
                DATE_FORMAT(fe.date, '%Y-%m-%d') = (  
                    SELECT MAX(DATE_FORMAT(date, '%Y-%m-%d'))   
                    FROM t_forex_exch
                    WHERE DATE_FORMAT(date, '%Y-%m') = DATE_FORMAT(fe.date, '%Y-%m'))
                AND fe.status_attachment = '0'"""

    sql = part1 + part2 + part3


    datas = []
    try:
        cursor.execute(sql)
        datas = cursor.fetchall()
        # print(datas)
    except mysql.connector.Error as err:
        print("Failed to Select forex info from table t_forex_exch: {}".format(err))

    if datas:
        data = [i for i in datas][0]

        last_date = data[0]
        dateobj = datetime.datetime.strptime(last_date, '%Y-%m-%d')
        cur_year = dateobj.year
        cur_month = dateobj.month

        if cur_month == 12:
            next_month = '1'
            cur_year = cur_year + 1
        else:
            next_month = str(int(cur_month) + 1)

        yymm =  f'{cur_year}/{next_month}/1'
        # print(yymm)

        forex_data_list = []
        forex_data_list.append(yymm)
        for j in data[1:]:
            j = float(j)
            forex_data_list.append(j)

        print(f'{yymm} : {forex_data_list}')

#####################################################################################################

    # read the excel file, and prepare to save the data into excel file
    # create DataFrame
        dfs = pd.DataFrame(np.array([forex_data_list]), columns=[
            '月份',
            '1USD→RMB',
            '1RMB→USD',
            '1HKD→RMB',
            '1RMB→HKD',
            '1JPY→RMB',
            '1RMB→JPY',
            '1HKD→USD',
            '1USD→HKD',
            '1JPY→USD',
            '1USD→JPY',
            '1EUR→RMB',
            '1RMB→EUR'])

        updating_sql = "UPDATE t_forex_exch SET status_attachment = '1' WHERE date = " + f"'{last_date}'"

        # setting the font and alignment
        font = Font(name='宋体', size=10)
        alignment = Alignment(horizontal='right')

        try:
            df = pd.read_excel(r'\\192.1.8.247\mis\forex_data\forex_table.xlsx', skiprows=1)
            row = df.shape[0]
            dfs = pd.concat([pd.DataFrame(columns=df.columns), dfs], ignore_index=True)

            book = openpyxl.load_workbook(r'\\192.1.8.247\mis\forex_data\forex_table.xlsx')

            with pd.ExcelWriter(r'\\192.1.8.247\mis\forex_data\forex_table.xlsx') as writer:
                writer.book = book
                dfs.to_excel(writer, sheet_name='Sheet1', startrow=row+2, index=False, header=False)

                # obtain the worksheet
                worksheet = writer.book['Sheet1']

                # set the font and alignment
                for row in range(3, row+5):
                    for col in range(1, len(df.columns)+1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = font
                        cell.alignment = alignment

            print(f"{date} {checktime}: Successfully insert the forex data into excel table 'forex_table'!~")

            cursor.execute(updating_sql)
            conn.commit()
            print(f"{date} {checktime}: Successfully update the status_attachment column of the table 't_forex_exch'!~")
        except Exception as e:
            print(e)

    return



### main function
def main():
    ### set the headless option to the chrome
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')

    ### set the initial url about "People Republic of China Bank"
    url = "http://www.pbc.gov.cn/zhengcehuobisi/125207/125217/125925/index.html"
    old_content = ""

    flag = False

    while True:

        ### get the local time
        check_time = time.strftime("%H:%M", time.localtime())
        ### get the date today
        today = datetime.date.today()

##########################################################################################################
        ### get the date of the end of this month, the start of this month
        this_month_end = datetime.date(
            today.year, today.month,
            calendar.monthrange(today.year, today.month)[1])

        # now_date = this_month_end - datetime.timedelta(days=11)

        ### only active once during today is the date end of this month
        if today == this_month_end and check_time >= "13:30" and check_time <= "14:20":
            if not flag:
                excel_forex_save(today, check_time)
                flag = True
            else:
                print(f"{today} {check_time}: It is already updated the data into 'forex_table' excel file")
                time.sleep(600)
        else:
            flag = False
        # break
##########################################################################################################

        if today not in holidays.country_holidays('CN') and today.weekday() < 5 and check_data_in_db(today):
            if check_time >= "09:20" and check_time <= "10:10":

                ### catch the content of html
                new_content = get_html_url(options, url)

                ### check the update
                if check_update(old_content, new_content) and check_href_in_db(new_content):
                    print(f"Webpage already updated! Great~")
                    old_content = new_content
                    print(f'{today} {check_time}: {new_content}')

                    driver = webdriver.Chrome(options=options)

                    c_to_r, r_to_c = data_obtain(driver, new_content)
                    process_save(c_to_r, r_to_c, today, new_content)
                    
                    driver.quit()

                    print(f'Forex Information crawled successfully, Waiting for the next update......')
                    time.sleep(5)

                else:
                    print(f'{today} {check_time}: Waiting for the webpage updating......')
                    time.sleep(180)
            else:
                print(f'{today} {check_time}: system sleeping...')
                time.sleep(1200)
        else:
            print(f"{today} {check_time}: Oh, System not in working time / todays' date already in DB")
            time.sleep(1800)


if __name__ == '__main__':
    main()
    
